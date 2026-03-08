import openpyxl
import json
import os
import re

# Use absolute path or correct relative path
file_path = os.path.join(os.path.dirname(__file__), '../東京3_26 - 3_31.xlsx')

def clean_value(val):
    if val is None:
        return ""
    return str(val).strip()

def extract_url_from_cell(cell):
    """從儲存格提取連結"""
    if cell.hyperlink:
        return cell.hyperlink.target
    # 如果沒有超連結，檢查文字中是否包含 URL
    value = clean_value(cell.value)
    url_match = re.search(r'(https?://[^\s]+)', value)
    if url_match:
        return url_match.group(1)
    return None

try:
    print(f"Reading file from: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb.active
    
    print(f"Sheet: {sheet.title}")
    print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")
    
    # 檢查最左邊的列（第1-5列）
    links = []
    notes_lines = []
    
    # 檢查前幾列
    for col in range(1, min(6, sheet.max_column + 1)):
        print(f"\n=== Column {col} ===")
        for row in range(1, min(sheet.max_row + 1, 50)):  # 檢查前50行
            cell = sheet.cell(row=row, column=col)
            value = clean_value(cell.value)
            url = extract_url_from_cell(cell)
            
            if url:
                print(f"  Row {row}: {value[:50] if value else '(empty)'} -> {url}")
                if url not in links:
                    links.append(url)
                    if value and value != url:
                        notes_lines.append(f"{value}: {url}")
                    else:
                        notes_lines.append(url)
            elif value and len(value) > 0:
                # 檢查文字中是否包含 URL
                url_match = re.search(r'(https?://[^\s]+)', value)
                if url_match:
                    url = url_match.group(1)
                    if url not in links:
                        links.append(url)
                        notes_lines.append(value)
                elif not value.startswith('Day') and not value.startswith('時間') and '行程' not in value:
                    # 可能是備註文字
                    if value not in notes_lines and len(value) > 3:
                        notes_lines.append(value)
    
    # 組合備註內容
    notes_content = '\n'.join(notes_lines)
    
    print(f"\n=== 提取結果 ===")
    print(f"找到 {len(links)} 個連結")
    print(f"備註內容:\n{notes_content[:200]}...")
    
    # 讀取現有的 trips 數據
    trips_file = os.path.join(os.path.dirname(__file__), 'src', 'data.json')
    trips_storage_file = os.path.join(os.path.dirname(__file__), '..', 'tokyo-trip', 'src', 'pages', 'HomePage.jsx')
    
    # 嘗試讀取 localStorage 中的 trips（如果有的話）
    # 或者直接更新預設旅程的 notes
    print(f"\n備註內容已準備好，可以手動添加到旅程的備註區域")
    print(f"\n備註內容（完整）:\n{notes_content}")
    
    # 保存到文件以便查看
    output_file = os.path.join(os.path.dirname(__file__), 'extracted_notes.txt')
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(notes_content)
    
    print(f"\n備註內容已保存到: {output_file}")

except Exception as e:
    print(f"Error processing Excel file: {e}")
    import traceback
    traceback.print_exc()
