import openpyxl
import json
import os
import re
from datetime import datetime

# Use absolute path or correct relative path
file_path = os.path.join(os.path.dirname(__file__), '../東京3_26 - 3_31.xlsx')

def clean_value(val):
    if val is None:
        return ""
    return str(val).strip()

def format_time(time_str):
    """格式化時間字串為標準格式 HH:MM"""
    if not time_str:
        return ""
    
    time_str = time_str.strip()
    
    def format_single_time(t):
        if not t:
            return ""
        t = t.strip()
        
        # 如果已經是標準格式 (HH:MM)，直接返回並補零
        if ':' in t:
            parts = t.split(':')
            hour = parts[0].zfill(2)
            minute = parts[1].zfill(2) if len(parts) > 1 else '00'
            return f"{hour}:{minute}"
        
        # 處理 "5.30" 或 "12." 格式
        if '.' in t:
            parts = t.split('.')
            hour = parts[0].zfill(2)
            minute = parts[1].zfill(2) if len(parts) > 1 and parts[1] else '00'
            return f"{hour}:{minute}"
        
        # 如果只有數字，假設是小時
        if t.isdigit():
            return f"{t.zfill(2)}:00"
        
        return t
    
    # 處理範圍時間 "開始 ~ 結束"
    if '~' in time_str:
        parts = time_str.split('~')
        start = format_single_time(parts[0])
        end = format_single_time(parts[1]) if len(parts) > 1 else ''
        return f"{start} ~ {end}" if end else f"{start} ~"
    
    # 單一時間
    return format_single_time(time_str)

def format_date(date_str):
    """格式化日期字串為標準格式"""
    if not date_str:
        return ""
    
    # 嘗試解析 "Day 1 3/26 (四)" 或 "Day 1 3/27 (五)" 等格式
    # 提取日期部分
    match = re.search(r'(\d{1,2})/(\d{1,2})\s*\(([一二三四五六日])\)', date_str)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        weekday = match.group(3)
        
        # 轉換星期為完整格式
        weekday_map = {
            '一': '週一', '二': '週二', '三': '週三', '四': '週四',
            '五': '週五', '六': '週六', '日': '週日'
        }
        weekday_full = weekday_map.get(weekday, weekday)
        
        # 格式化為 "3月26日 (週四)" 或 "3/26 (週四)"
        return f"{month}月{day}日 ({weekday_full})"
    
    # 如果無法解析，嘗試其他格式
    # 例如 "3/26" 或 "3-26"
    match = re.search(r'(\d{1,2})[/-](\d{1,2})', date_str)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        return f"{month}月{day}日"
    
    # 如果都無法解析，返回原始字串但清理一下
    return date_str.replace('Day ', '').strip()

try:
    print(f"Reading file from: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb.active
    
    days = []
    
    # Column indices (1-based)
    # Day 1: 6 (Time), 7 (Activity), 8 (Cost TWD), 9 (Cost JPY)
    # Stride: 6
    day_starts = [6, 12, 18, 24, 30, 36]
    
    for start_col in day_starts:
        # Check if column exists
        if start_col > sheet.max_column:
            break
            
        # Get date from header (Row 1, Activity column)
        date_cell = sheet.cell(row=1, column=start_col + 1)
        date_header_raw = clean_value(date_cell.value)
        
        if not date_header_raw:
            continue
        
        # Format the date
        date_header = format_date(date_header_raw)
            
        day_events = []
        
        # Iterate rows
        for row in range(2, sheet.max_row + 1):
            time_cell = sheet.cell(row=row, column=start_col)
            activity_cell = sheet.cell(row=row, column=start_col + 1)
            cost_twd_cell = sheet.cell(row=row, column=start_col + 2)
            cost_jpy_cell = sheet.cell(row=row, column=start_col + 3)
            
            time_val_raw = clean_value(time_cell.value)
            activity_val = clean_value(activity_cell.value)
            
            # 格式化時間
            time_val = format_time(time_val_raw)
            
            # Skip empty rows or header-like rows
            if not time_val and not activity_val:
                continue
            if activity_val == "行程" or time_val == "時間 (預估)":
                continue
            if "Day" in time_val or "Day" in activity_val: # Skip if it reads a header row by mistake
                continue
                
            # Extract hyperlink
            link = ""
            if activity_cell.hyperlink:
                link = activity_cell.hyperlink.target
            
            # Get costs
            cost_twd = clean_value(cost_twd_cell.value)
            cost_jpy = clean_value(cost_jpy_cell.value)
            
            day_events.append({
                "time": time_val,
                "activity": activity_val,
                "cost_twd": cost_twd,
                "cost_jpy": cost_jpy,
                "link": link
            })
            
        if day_events:
            days.append({
                "date": date_header,
                "events": day_events
            })
            
    # Save to src/data.json
    output_path = os.path.join(os.path.dirname(__file__), 'src/data.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(days, f, ensure_ascii=False, indent=2)
        
    print(f"Successfully converted Excel to {output_path}")
    
    # 提取最左邊列的備註和連結
    print("\n提取備註與連結...")
    notes_lines = []
    seen_urls = set()
    
    def extract_url_from_cell(cell):
        """從儲存格提取連結"""
        if cell.hyperlink:
            return cell.hyperlink.target
        value = clean_value(cell.value)
        if value:
            match = re.search(r'(https?://[^\s\n]+)', value)
            if match:
                return match.group(1)
        return None
    
    # 檢查最左邊的列（1-5列）
    for col in range(1, 6):
        for row in range(1, min(200, sheet.max_row + 1)):
            cell = sheet.cell(row=row, column=col)
            value = clean_value(cell.value)
            url = extract_url_from_cell(cell)
            
            if url and url not in seen_urls:
                seen_urls.add(url)
                if value and value != url and len(value) > 5:
                    notes_lines.append(f"{value}\n{url}")
                else:
                    notes_lines.append(url)
            elif value and len(value) > 3:
                url_match = re.search(r'(https?://[^\s\n]+)', value)
                if url_match:
                    url = url_match.group(1)
                    if url not in seen_urls:
                        seen_urls.add(url)
                        notes_lines.append(value)
                elif not any(k in value for k in ['Day', '時間', '行程', '活動', 'Cost', 'JPY', 'TWD', '預估', 'Time', 'Activity']):
                    if value not in notes_lines:
                        notes_lines.append(value)
    
    notes_content = '\n'.join(notes_lines)
    
    # 保存備註到 JSON 文件
    notes_output_path = os.path.join(os.path.dirname(__file__), 'src', 'default_notes.json')
    notes_data = {
        'notes': notes_content,
        'links': list(seen_urls)
    }
    
    with open(notes_output_path, 'w', encoding='utf-8') as f:
        json.dump(notes_data, f, ensure_ascii=False, indent=2)
    
    print(f"✅ 已提取 {len(seen_urls)} 個連結")
    print(f"✅ 備註已保存到: {notes_output_path}")

except Exception as e:
    print(f"Error processing Excel file: {e}")
    import traceback
    traceback.print_exc()