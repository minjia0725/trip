import openpyxl
import json
import os
import re

# 獲取當前腳本所在目錄的父目錄
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
excel_path = os.path.join(parent_dir, '東京3_26 - 3_31.xlsx')

def clean_value(val):
    if val is None:
        return ""
    return str(val).strip()

def extract_url_from_cell(cell):
    """從儲存格提取連結"""
    if cell.hyperlink:
        return cell.hyperlink.target
    value = clean_value(cell.value)
    if value:
        url_match = re.search(r'(https?://[^\s\n]+)', value)
        if url_match:
            return url_match.group(1)
    return None

try:
    print(f"讀取 Excel 文件: {excel_path}")
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        # 嘗試另一個路徑
        excel_path = os.path.join(current_dir, '..', '東京3_26 - 3_31.xlsx')
        if not os.path.exists(excel_path):
            print(f"文件也不存在: {excel_path}")
            exit(1)
    
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    sheet = wb.active
    
    print(f"工作表: {sheet.title}")
    
    # 收集所有連結和文字（從最左邊的列）
    notes_lines = []
    seen_urls = set()
    
    # 檢查第1-5列（最左邊的列）
    for col in range(1, min(6, sheet.max_column + 1)):
        for row in range(1, min(sheet.max_row + 1, 200)):  # 檢查前200行
            cell = sheet.cell(row=row, column=col)
            value = clean_value(cell.value)
            url = extract_url_from_cell(cell)
            
            if url and url not in seen_urls:
                seen_urls.add(url)
                if value and value != url and len(value) > 5:
                    notes_lines.append(f"{value}\n{url}")
                else:
                    notes_lines.append(url)
            elif value and len(value) > 0:
                # 檢查文字中是否包含 URL
                url_match = re.search(r'(https?://[^\s\n]+)', value)
                if url_match:
                    url = url_match.group(1)
                    if url not in seen_urls:
                        seen_urls.add(url)
                        notes_lines.append(value)
                elif not any(keyword in value for keyword in ['Day', '時間', '行程', '活動', 'Cost', 'JPY', 'TWD', '預估']):
                    if len(value) > 3 and value not in notes_lines:
                        notes_lines.append(value)
    
    # 組合備註內容
    notes_content = '\n'.join(notes_lines)
    
    print(f"\n=== 提取結果 ===")
    print(f"找到 {len(seen_urls)} 個連結")
    print(f"備註行數: {len(notes_lines)}")
    print(f"\n備註內容:\n{notes_content}")
    
    # 更新 HomePage.jsx 中的預設 notes
    homepage_path = os.path.join(current_dir, 'src', 'pages', 'HomePage.jsx')
    if os.path.exists(homepage_path):
        with open(homepage_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 轉義 notes_content 中的特殊字符
        escaped_notes = notes_content.replace('\\', '\\\\').replace('`', '\\`').replace('${', '\\${')
        
        # 更新 notes 字段
        pattern = r"(notes:\s*)''"
        replacement = f"notes: `{escaped_notes}`"
        new_content = re.sub(pattern, replacement, content)
        
        if new_content != content:
            with open(homepage_path, 'w', encoding='utf-8') as f:
                f.write(new_content)
            print(f"\n已更新 HomePage.jsx 中的預設 notes")
        else:
            print(f"\n未找到 notes: '' 模式，請手動更新")
            print(f"\n請將以下內容複製到 HomePage.jsx 的 defaultTrip.notes 中：")
            print(f"\n{notes_content}")
    else:
        print(f"\n未找到 HomePage.jsx，請手動更新")
        print(f"\n請將以下內容複製到旅程的備註區域：")
        print(f"\n{notes_content}")

except Exception as e:
    print(f"錯誤: {e}")
    import traceback
    traceback.print_exc()
