# -*- coding: utf-8 -*-
import openpyxl
import json
import os
import re

# 使用與 extract_data_v2.py 相同的路徑邏輯
file_path = os.path.join(os.path.dirname(__file__), '../東京3_26 - 3_31.xlsx')

def clean_value(val):
    if val is None:
        return ""
    return str(val).strip()

def extract_url(cell):
    """提取連結"""
    if cell.hyperlink:
        return cell.hyperlink.target
    value = clean_value(cell.value)
    if value:
        match = re.search(r'(https?://[^\s\n]+)', value)
        if match:
            return match.group(1)
    return None

try:
    print(f"讀取 Excel 文件: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb.active
    
    print(f"工作表: {sheet.title}")
    print(f"最大行數: {sheet.max_row}, 最大列數: {sheet.max_column}")
    
    notes_lines = []
    seen_urls = set()
    
    # 檢查最左邊的列（1-5列）
    for col in range(1, 6):
        print(f"\n檢查第 {col} 列...")
        for row in range(1, min(200, sheet.max_row + 1)):
            cell = sheet.cell(row=row, column=col)
            value = clean_value(cell.value)
            url = extract_url(cell)
            
            if url and url not in seen_urls:
                seen_urls.add(url)
                if value and value != url and len(value) > 5:
                    notes_lines.append(f"{value}\n{url}")
                else:
                    notes_lines.append(url)
            elif value and len(value) > 3:
                # 檢查是否包含 URL
                url_match = re.search(r'(https?://[^\s\n]+)', value)
                if url_match:
                    url = url_match.group(1)
                    if url not in seen_urls:
                        seen_urls.add(url)
                        notes_lines.append(value)
                elif not any(k in value for k in ['Day', '時間', '行程', '活動', 'Cost', 'JPY', 'TWD', '預估', 'Time', 'Activity']):
                    if value not in notes_lines:
                        notes_lines.append(value)
    
    notes_content = '\n'.join(notes_lines)
    
    print(f"\n=== 提取結果 ===")
    print(f"找到 {len(seen_urls)} 個連結")
    print(f"備註行數: {len(notes_lines)}")
    print(f"\n備註內容:\n{notes_content}")
    
    # 更新 HomePage.jsx
    homepage_path = os.path.join(os.path.dirname(__file__), 'src', 'pages', 'HomePage.jsx')
    if os.path.exists(homepage_path):
        with open(homepage_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 轉義特殊字符
        escaped_notes = notes_content.replace('\\', '\\\\').replace('`', '\\`').replace('${', '\\${').replace('$', '\\$')
        
        # 更新 notes 字段（使用模板字符串）
        pattern = r"(notes:\s*)''"
        replacement = f"notes: `{escaped_notes}`"
        new_content = re.sub(pattern, replacement, content)
        
        if new_content != content:
            with open(homepage_path, 'w', encoding='utf-8') as f:
                f.write(new_content)
            print(f"\n✅ 已更新 HomePage.jsx 中的預設 notes")
        else:
            print(f"\n⚠️ 未找到 notes: '' 模式")
            print(f"\n請手動將以下內容複製到 HomePage.jsx 的 defaultTrip.notes 中：")
            print(f"\n{notes_content}")
    else:
        print(f"\n⚠️ 未找到 HomePage.jsx: {homepage_path}")
        print(f"\n請手動將以下內容複製到旅程的備註區域：")
        print(f"\n{notes_content}")

except Exception as e:
    print(f"錯誤: {e}")
    import traceback
    traceback.print_exc()
