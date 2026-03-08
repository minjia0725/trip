import openpyxl
import os

file_path = os.path.join(os.path.dirname(__file__), '../東京3_26 - 3_31.xlsx')

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb.active
    
    day_starts = [6, 12, 18, 24, 30, 36]
    
    print("檢查 Excel 中的時間格式：\n")
    
    for day_idx, start_col in enumerate(day_starts):
        if start_col > sheet.max_column:
            break
            
        date_cell = sheet.cell(row=1, column=start_col + 1)
        date_header = str(date_cell.value).strip() if date_cell.value else ""
        
        if not date_header:
            continue
            
        print(f"\n=== Day {day_idx + 1}: {date_header} ===")
        
        # 檢查前 10 行的時間
        for row in range(2, min(12, sheet.max_row + 1)):
            time_cell = sheet.cell(row=row, column=start_col)
            activity_cell = sheet.cell(row=row, column=start_col + 1)
            
            time_val = str(time_cell.value).strip() if time_cell.value else ""
            activity_val = str(activity_cell.value).strip() if activity_cell.value else ""
            
            if time_val or activity_val:
                if activity_val != "行程" and time_val != "時間 (預估)":
                    if "Day" not in time_val and "Day" not in activity_val:
                        print(f"  行 {row}: 時間='{time_val}' | 活動='{activity_val[:30]}...'")
                        
except Exception as e:
    print(f"錯誤: {e}")
