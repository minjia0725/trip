# -*- coding: utf-8 -*-
import openpyxl
import json
import os
import re

# 使用絕對路徑
base_dir = r'C:\Users\aa813\OneDrive\桌面\東京'
excel_file = os.path.join(base_dir, '東京3_26 - 3_31.xlsx')
output_file = os.path.join(base_dir, 'tokyo-trip', 'src', 'default_notes.json')

print(f"Excel 文件: {excel_file}")
print(f"輸出文件: {output_file}")

def clean_value(val):
    if val is None:
        return ""
    return str(val).strip()

def extract_url(cell):
    if cell.hyperlink:
        return cell.hyperlink.target
    value = clean_value(cell.value)
    if value:
        match = re.search(r'(https?://[^\s\n]+)', value)
        if match:
            return match.group(1)
    return None

try:
    wb = openpyxl.load_workbook(excel_file, data_only=False)
    sheet = wb.active
    
    notes_lines = []
    seen_urls = set()
    
    for col in range(1, 6):
        for row in range(1, min(200, sheet.max_row + 1)):
            cell = sheet.cell(row=row, column=col)
            value = clean_value(cell.value)
            url = extract_url(cell)
            
            if url and url not in seen_urls:
                seen_urls.add(url)
                if value and value != url and len(value) > 5:
                    notes_lines.append(f"{value}\n{url}")
                else:
                    notes_lines.append(url)
            elif value and len(value) > 3:
                url_match = re.search(r'(https?://[^\s\n]+)', value)
                if url_match:
                    url = url_match.group(1)
                    if url not in seen_urls:
                        seen_urls.add(url)
                        notes_lines.append(value)
                elif not any(k in value for k in ['Day', '時間', '行程', '活動', 'Cost', 'JPY', 'TWD', '預估', 'Time', 'Activity']):
                    if value not in notes_lines:
                        notes_lines.append(value)
    
    notes_content = '\n'.join(notes_lines)
    
    notes_data = {
        'notes': notes_content,
        'links': list(seen_urls)
    }
    
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(notes_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ 已提取 {len(seen_urls)} 個連結")
    print(f"✅ 已保存到: {output_file}")
    print(f"\n備註內容預覽:\n{notes_content[:300]}...")

except Exception as e:
    print(f"錯誤: {e}")
    import traceback
    traceback.print_exc()
