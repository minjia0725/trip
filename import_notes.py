import openpyxl
import json
import os
import re

# 獲取當前腳本所在目錄
script_dir = os.path.dirname(os.path.abspath(__file__))
# Excel 文件路徑
excel_path = os.path.join(script_dir, '..', '東京3_26 - 3_31.xlsx')
# 輸出 JSON 文件路徑
output_path = os.path.join(script_dir, 'src', 'notes.json')

def clean_value(val):
    if val is None:
        return ""
    return str(val).strip()

def extract_url_from_cell(cell):
    """從儲存格提取連結"""
    if cell.hyperlink:
        return cell.hyperlink.target
    # 如果沒有超連結，檢查文字中是否包含 URL
    value = clean_value(cell.value)
    if value:
        url_match = re.search(r'(https?://[^\s\n]+)', value)
        if url_match:
            return url_match.group(1)
    return None

try:
    print(f"讀取 Excel 文件: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    sheet = wb.active
    
    print(f"工作表: {sheet.title}")
    print(f"最大行數: {sheet.max_row}, 最大列數: {sheet.max_column}")
    
    # 收集所有連結和文字
    links = []
    notes_lines = []
    seen_urls = set()
    
    # 檢查最左邊的列（第1-5列）
    for col in range(1, min(6, sheet.max_column + 1)):
        print(f"\n檢查第 {col} 列...")
        for row in range(1, min(sheet.max_row + 1, 100)):  # 檢查前100行
            cell = sheet.cell(row=row, column=col)
            value = clean_value(cell.value)
            url = extract_url_from_cell(cell)
            
            if url and url not in seen_urls:
                seen_urls.add(url)
                links.append(url)
                if value and value != url and len(value) > 5:
                    # 如果有文字且不是純 URL，保留文字和連結
                    notes_lines.append(f"{value}\n{url}")
                else:
                    notes_lines.append(url)
            elif value and len(value) > 0:
                # 檢查文字中是否包含 URL
                url_match = re.search(r'(https?://[^\s\n]+)', value)
                if url_match:
                    url = url_match.group(1)
                    if url not in seen_urls:
                        seen_urls.add(url)
                        links.append(url)
                        notes_lines.append(value)
                elif not any(keyword in value for keyword in ['Day', '時間', '行程', '活動', 'Cost', 'JPY', 'TWD']):
                    # 可能是備註文字，排除標題行
                    if len(value) > 3 and value not in notes_lines:
                        notes_lines.append(value)
    
    # 組合備註內容
    notes_content = '\n'.join(notes_lines)
    
    print(f"\n=== 提取結果 ===")
    print(f"找到 {len(links)} 個連結")
    print(f"備註行數: {len(notes_lines)}")
    print(f"\n備註內容預覽:\n{notes_content[:300]}...")
    
    # 保存到 JSON 文件
    notes_data = {
        'notes': notes_content,
        'links': links,
        'extracted_at': str(os.path.getmtime(excel_path))
    }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(notes_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n備註內容已保存到: {output_path}")
    print(f"\n完整備註內容:\n{notes_content}")
    
    # 嘗試更新 localStorage 中的 trips
    # 讀取現有的 trips（如果有的話）
    print(f"\n請手動將以下內容複製到旅程的備註區域，或運行更新腳本自動更新。")

except Exception as e:
    print(f"錯誤: {e}")
    import traceback
    traceback.print_exc()
